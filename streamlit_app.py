import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import io
import os
import re
from datetime import datetime

# ============================================================================
# CẤU HÌNH & HÀM HỖ TRỢ
# ============================================================================
st.set_page_config(page_title="Export Tool Web", layout="wide")

def clean_filename(text):
    return re.sub(r'[\\/*?:"<>|]', "", text)

def try_parse_number(value):
    if not isinstance(value, str): return value
    val = value.strip()
    if not val: return value
    if val.startswith("="): return val
    try:
        # Xử lý dấu phẩy
        return float(val.replace(",", ""))
    except:
        return value

# Hàm load company list
def get_company_list(wb):
    companies = []
    if "Data Base" in wb.sheetnames:
        ws = wb["Data Base"]
        # Giả sử cột Company là cột B (index 1)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) > 1 and row[1]:
                companies.append(str(row[1]))
    return sorted(list(set(companies)))

# Hàm mapping dữ liệu (Logic gốc của bạn)
def map_data(wb, company, pol_pod):
    # 1. Map Form Data
    ws_db = wb["Data Base"]
    ws_pol = wb["POL-POD"]
    
    form_data = []
    
    # Tìm Company trong Data Base
    com_row = ["1", company, "", "", "", "", "", ""]
    for row in ws_db.iter_rows(min_row=2, values_only=True):
        if len(row) > 1 and str(row[1]) == company:
            # Excel: A=0, B=1(Name), C=2(Desc), D=3(Rate), E=4(Unit), F=5(Cur), G=6(VAT), H=7(PP)
            com_row = [
                "1", 
                str(row[2] or ""), str(row[3] or ""), 
                str(row[4] or ""), str(row[5] or ""), 
                "", # Amount trống
                str(row[6] or ""), str(row[7] or "")
            ]
            break
    form_data.append(com_row)
    
    # Tìm POL-POD
    pod_key = pol_pod.split("-")[-1] if "-" in pol_pod else ""
    group_1 = ["BKK", "SHA", "HKG", "SIN", "NGB", "JKT"]
    group_2 = ["INC", "CGP"]
    
    rows_to_get = []
    if pod_key in group_1: rows_to_get = range(3, 9) # Excel row index
    elif pod_key in group_2: rows_to_get = range(12, 17)
    
    # Đọc dữ liệu POL-POD (Openpyxl row bắt đầu từ 1)
    idx_cnt = 2
    for r_idx in (rows_to_get or []):
        # Lấy row từ sheet (index 1-based)
        # Giả sử cột dữ liệu giống Data Base bắt đầu từ cột B
        desc = ws_pol.cell(row=r_idx, column=2).value
        rate = ws_pol.cell(row=r_idx, column=3).value
        unit = ws_pol.cell(row=r_idx, column=4).value
        curr = ws_pol.cell(row=r_idx, column=5).value
        vat  = ws_pol.cell(row=r_idx, column=6).value
        pp   = ws_pol.cell(row=r_idx, column=7).value
        
        row_data = [
            str(idx_cnt),
            str(desc or ""), str(rate or ""), 
            str(unit or ""), str(curr or ""), 
            "", # Amount
            str(vat or ""), str(pp or "")
        ]
        form_data.append(row_data)
        idx_cnt += 1
        
    return form_data, pod_key

# ============================================================================
# GIAO DIỆN CHÍNH
# ============================================================================
st.title("🚢 EXPORT TOOL (Web Version)")

# 1. Upload Template
uploaded_file = st.file_uploader("Upload file Template.xlsx", type="xlsx")

if uploaded_file:
    # Load workbook vào session state để không phải load lại
    try:
        wb = load_workbook(uploaded_file, data_only=True) # Để đọc dữ liệu map
        # Cần load thêm bản không data_only để giữ công thức khi save
        wb_formula = load_workbook(uploaded_file, data_only=False) 
        
        companies = get_company_list(wb)
    except Exception as e:
        st.error(f"Lỗi đọc file: {e}")
        st.stop()

    # --- KHUNG NHẬP LIỆU ---
    with st.container():
        st.subheader("Thông tin lô hàng")
        col1, col2 = st.columns(2)
        
        with col1:
            job_file = st.text_input("Job file (Tự động viết hoa):", key="job_input").upper()
            
            # Logic tự động điền POL-POD
            pol_pod_default = ""
            if len(job_file) >= 6:
                pol_pod_default = f"{job_file[0:3]}-{job_file[3:6]}"
            
            pol_pod = st.text_input("POL/POD:", value=pol_pod_default, disabled=True)
            commodity = st.text_input("Commodity:")
            volume = st.text_input("E. Volume:")
            etd = st.date_input("ETD / ETA:", format="DD/MM/YYYY")

        with col2:
            company = st.selectbox("To (Company):", options=companies)
            attn = st.text_input("Attn:")
            tel = st.text_input("Tel:")
            sop = st.text_input("SOP No.:")
            date_create = st.date_input("Date:", format="DD/MM/YYYY")

    # --- NÚT LOAD PREVIEW ---
    if st.button("LOAD & MAP TEMPLATE", type="primary"):
        # Map dữ liệu
        form_mapped, pod_name = map_data(wb, company, pol_pod)
        
        # Lưu vào session để hiển thị
        st.session_state['form_data'] = form_mapped
        st.session_state['pod_name'] = pod_name
        
        # Load dữ liệu POD Sheet để Preview
        if pod_name and pod_name in wb_formula.sheetnames:
            ws_pod = wb_formula[pod_name]
            # Đọc dữ liệu ra DataFrame
            data = list(ws_pod.values)
            if data:
                # Chuyển thành DataFrame để hiển thị và edit
                df = pd.DataFrame(data)
                # Chuyển đổi None thành ""
                df = df.fillna("")
                st.session_state['pod_df'] = df
            else:
                st.session_state['pod_df'] = pd.DataFrame()
        else:
             st.warning(f"Không tìm thấy sheet POD: {pod_name}")
             st.session_state['pod_df'] = pd.DataFrame()

    # --- KHUNG PREVIEW VÀ EDIT ---
    if 'form_data' in st.session_state:
        st.divider()
        tab1, tab2 = st.tabs(["Preview: Form", "Preview: POD Sheet"])
        
        with tab1:
            # Hiển thị bảng Form (chỉ xem)
            df_form = pd.DataFrame(st.session_state['form_data'], 
                                   columns=["No", "Description", "Rate", "Unit", "Cur", "Amount", "VAT", "PP/CC"])
            st.dataframe(df_form, use_container_width=True, hide_index=True)
            
        with tab2:
            st.info("💡 Lưu ý: Trên Web, công thức sẽ KHÔNG tự động tính toán ngay lập tức. Tuy nhiên khi bạn Xuất file Excel và mở trên máy tính, mọi công thức sẽ hoạt động bình thường.")
            
            if not st.session_state['pod_df'].empty:
                # Cho phép chỉnh sửa trực tiếp
                edited_df = st.data_editor(st.session_state['pod_df'], use_container_width=True, num_rows="dynamic", hide_index=True)
                st.session_state['edited_pod_df'] = edited_df

    # --- NÚT EXPORT ---
    st.divider()
    if 'edited_pod_df' in st.session_state:
        
        # Chuẩn bị tên file
        safe_job = clean_filename(job_file)
        safe_com = clean_filename(company)
        file_name = f"QUO {safe_job} {safe_com}.xlsx"
        
        # Logic Xuất File
        def convert_df_to_excel():
            output = io.BytesIO()
            # Load lại wb gốc để giữ format
            # Lưu ý: wb_formula đã load ở trên
            
            # 1. Update Sheet Form
            ws_form = wb_formula["Form"]
            # Clear dữ liệu cũ (A8:H20)
            for r in range(8, 21):
                for c in range(1, 9):
                    ws_form.cell(row=r, column=c).value = None
            
            # Ghi dữ liệu mới vào Form
            form_data = st.session_state['form_data']
            for i, row_data in enumerate(form_data):
                # row_data là list string
                for j, val in enumerate(row_data):
                    # i=0 -> Row 8
                    ws_form.cell(row=8+i, column=j+1).value = try_parse_number(val)

            # 2. Update Header Form
            date_str = date_create.strftime("%d-%b-%y")
            etd_str = etd.strftime("%d-%b-%y")
            
            replace_map = {
                "{{job_id}}": job_file, "{{com}}": commodity,
                "{{volume}}": volume, "{{ETD}}": etd_str,
                "{{company}}": company,
                "{{pol_pod}}": pol_pod, "{{form-to}}": pol_pod,
                "{{attn}}": attn, "{{tel}}": tel,
                "{{sop}}": sop, "{{date}}": date_str
            }
            
            # Quét header 7 dòng đầu
            for r in range(1, 8):
                for c in range(1, 11):
                    cell = ws_form.cell(row=r, column=c)
                    if isinstance(cell.value, str):
                        for k, v in replace_map.items():
                            if k in cell.value:
                                cell.value = cell.value.replace(k, v)
                                
            # Footer Date
            ws_form["F40"].value = date_str
            ws_form["H40"].value = date_str
            
            # 3. Update POD Sheet
            pod_name = st.session_state['pod_name']
            if pod_name:
                if pod_name not in wb_formula.sheetnames:
                    wb_formula.create_sheet(pod_name)
                
                ws_target = wb_formula[pod_name]
                # Lấy dữ liệu đã edit từ Data Editor
                df_final = st.session_state['edited_pod_df']
                
                # Ghi đè vào Excel
                for r_idx, row in df_final.iterrows():
                    for c_idx, val in enumerate(row):
                        # Ghi vào (r_idx+1, c_idx+1)
                        cell = ws_target.cell(row=r_idx+1, column=c_idx+1)
                        # Nếu là công thức thì giữ nguyên (nếu user không sửa)
                        # Nhưng data editor trả về string, nên ta cứ ghi đè
                        # Cẩn thận: Nếu cell cũ là công thức mà user không sửa, df có thể chứa kết quả cũ hoặc string
                        
                        # Logic đơn giản: Ghi giá trị, convert số nếu được
                        cell.value = try_parse_number(val)
                
                # Xóa các sheet thừa
                sheets_to_keep = ["Form", pod_name]
                for s in wb_formula.sheetnames:
                    if s not in sheets_to_keep:
                        del wb_formula[s]

            wb_formula.save(output)
            return output.getvalue()

        # Tạo nút download
        excel_data = convert_df_to_excel()
        st.download_button(
            label="📥 EXPORT TO EXCEL",
            data=excel_data,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

else:
    st.info("Vui lòng upload file Template.xlsx để bắt đầu.")